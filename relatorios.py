import pandas as pd
import numpy as np
from datetime import datetime, date, timedelta
import os
from openpyxl import Workbook
from openpyxl.styles import Font
from fpdf import FPDF
from supabase_client import QueryBuilder, select_all
from db_compat import get_all_users
import db_compat as utils

def get_available_years():
    """
    Busca no banco de dados todos os anos únicos em que processos foram atribuídos.
    """
    try:
        # Fetch all processes with data_atribuicao_servidor
        all_processes = select_all("processos")
        years = set()
        for p in all_processes:
            data_str = p.get('data_atribuicao_servidor')
            if data_str:
                try:
                    if isinstance(data_str, str):
                        year = int(data_str[:4])
                    else:
                        year = data_str.year
                    years.add(year)
                except:
                    pass
        # Ordena os anos em ordem decrescente para mostrar os mais recentes primeiro
        return sorted(list(years), reverse=True)
    except Exception as e:
        print(f"Erro ao buscar anos disponíveis para relatório: {e}")
        return [] # Retorna uma lista vazia em caso de erro


def sanitize_text(text):
    if text is None: return ""
    return str(text).encode('latin-1', 'replace').decode('latin-1')

class PDF(FPDF):
    def header(self):
        self.set_font('Arial', 'B', 15)
        try:
            self.image('logo_mpcsc.jpg', 10, 8, 33)
        except Exception:
            self.cell(0, 10, 'MPC/SC', 0, 0, 'L')
        self.cell(0, 10, sanitize_text('Relatório de Produtividade'), 0, 1, 'C')
        self.ln(10)

    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Pagina {self.page_no()}', 0, 0, 'C')

def gerar_relatorio_dashboard(df_processos: pd.DataFrame):
    pdf = PDF(orientation='L')
    pdf.add_page()
    
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(0, 10, sanitize_text('Lista de Processos'), 0, 1, 'L')

    if df_processos.empty:
        pdf.set_font('Arial', '', 12)
        pdf.cell(0, 10, 'Nenhum processo encontrado para os filtros selecionados.', 0, 1)
        return bytes(pdf.output(dest='S')) # Corrected line

    pdf.set_font('Arial', 'B', 8)
    line_height = pdf.font_size * 2.5
    col_widths = {
        'Nº Processo': 25,
        'Servidor': 45,
        'Produto': 65,
        'Status Geral': 25,
        'Data Final': 25
    }
    headers = list(col_widths.keys())

    for i, header in enumerate(headers):
        pdf.cell(col_widths[header], line_height, sanitize_text(header), 1, 0, 'C')
    pdf.ln()

    pdf.set_font('Arial', '', 8)
    for index, row in df_processos.iterrows():
        x_start = pdf.get_x()
        y_start = pdf.get_y()
        max_y = y_start
        pdf.set_xy(x_start, y_start)
        pdf.multi_cell(col_widths['Nº Processo'], line_height, sanitize_text(row.get('Nº Processo')), border=1, align='L')
        max_y = max(max_y, pdf.get_y())
        
        pdf.set_xy(x_start + col_widths['Nº Processo'], y_start)
        pdf.multi_cell(col_widths['Servidor'], line_height, sanitize_text(row.get('Servidor')), border=1, align='L')
        max_y = max(max_y, pdf.get_y())

        pdf.set_xy(x_start + col_widths['Nº Processo'] + col_widths['Servidor'], y_start)
        pdf.multi_cell(col_widths['Produto'], line_height, sanitize_text(row.get('Produto')), border=1, align='L')
        max_y = max(max_y, pdf.get_y())

        pdf.set_xy(x_start + col_widths['Nº Processo'] + col_widths['Servidor'] + col_widths['Produto'], y_start)
        pdf.multi_cell(col_widths['Status Geral'], line_height, sanitize_text(row.get('Status Geral')), border=1, align='C')
        max_y = max(max_y, pdf.get_y())
        
        pdf.set_xy(x_start + col_widths['Nº Processo'] + col_widths['Servidor'] + col_widths['Produto'] + col_widths['Status Geral'], y_start)
        pdf.multi_cell(col_widths['Data Final'], line_height, sanitize_text(row.get('Data Final')), border=1, align='C')
        max_y = max(max_y, pdf.get_y())

        pdf.set_y(max_y)
        
    return bytes(pdf.output(dest='S')) # Corrected line

def gerar_relatorio_detalhado(df_processos: pd.DataFrame):
    pdf = PDF(orientation='L')
    pdf.add_page()

    pdf.set_font('Arial', 'B', 12)
    pdf.cell(0, 10, sanitize_text('Relatório Detalhado de Processos'), 0, 1, 'C')
    pdf.ln()

    if df_processos.empty:
        pdf.set_font('Arial', '', 12)
        pdf.cell(0, 10, 'Nenhum processo encontrado para os filtros selecionados.', 0, 1)
        return bytes(pdf.output(dest='S')) # Corrected line

    pdf.set_font('Arial', 'B', 8)
    
    headers = [
        "Status", "Nº Processo", "Servidor",
        "Data Atribuição Servidor", "Prazo Servidor", "Data Conclusão Servidor",
        "Data Atribuição Chefe", "Prazo Chefe", "Data Conclusão"
    ]
    
    col_widths = {
        "Status": 20, "Nº Processo": 45, "Servidor": 40,
        "Data Atribuição Servidor": 38, "Prazo Servidor": 20, "Data Conclusão Servidor": 38,
        "Data Atribuição Chefe": 35, "Prazo Chefe": 20, "Data Conclusão": 25
    }

    for header in headers:
        pdf.cell(col_widths[header], 10, sanitize_text(header), 1, 0, 'C')
    pdf.ln()

    pdf.set_font('Arial', '', 8)
    
    for _, row in df_processos.iterrows():
        status = row.get('status_servidor', 'N/A')

        if status == "Atrasado": pdf.set_fill_color(255, 102, 102)
        elif status == "No Prazo": pdf.set_fill_color(102, 255, 102)
        elif status == "Concluído": pdf.set_fill_color(173, 216, 230)
        else: pdf.set_fill_color(255, 255, 255)

        pdf.cell(col_widths["Status"], 10, sanitize_text(status), 1, 0, 'C', fill=True)
        pdf.cell(col_widths["Nº Processo"], 10, sanitize_text(row.get('processo_numero')), 1)
        pdf.cell(col_widths["Servidor"], 10, sanitize_text(row.get('servidor_responsavel_nome')), 1)
        pdf.cell(col_widths["Data Atribuição Servidor"], 10, sanitize_text(str(row.get('data_atribuicao_servidor', ''))), 1)
        pdf.cell(col_widths["Prazo Servidor"], 10, sanitize_text(str(row.get('prazo_servidor_aplicado', ''))), 1)
        pdf.cell(col_widths["Data Conclusão Servidor"], 10, sanitize_text(str(row.get('data_conclusao_servidor', ''))), 1)
        pdf.cell(col_widths["Data Atribuição Chefe"], 10, sanitize_text(str(row.get('data_atribuicao_chefe', ''))), 1)
        pdf.cell(col_widths["Prazo Chefe"], 10, sanitize_text(str(row.get('prazo_chefe_aplicado', ''))), 1)
        pdf.cell(col_widths["Data Conclusão"], 10, sanitize_text(str(row.get('data_conclusao_chefe', ''))), 1)
        pdf.ln()
        
    return bytes(pdf.output(dest='S')) # Corrected line

def _get_user_hierarchy():
    """Get user hierarchy from Supabase API."""
    all_users = get_all_users()
    
    servidores = {u.get('id'): u.get('nome_completo') for u in all_users if u.get('perfil') == 'Servidor'}
    
    # For chefes, we need to fetch the relationships
    chefes = {}
    for u in all_users:
        if u.get('perfil') == 'Chefe de Gabinete':
            # Get servidores linked to this chefe via gabinete_servidores table
            servs = QueryBuilder("gabinete_servidores").eq("id_chefe", u.get('id')).execute()
            servidor_ids = [s.get('id_servidor') for s in servs]
            # Get procuradores linked via procurador_chefes table
            procs = QueryBuilder("procurador_chefes").eq("id_chefe", u.get('id')).execute()
            proc_ids = [p.get('id_procurador') for p in procs]
            chefes[u.get('id')] = {
                "nome": u.get('nome_completo'),
                "servidores": servidor_ids,
                "procuradores": proc_ids
            }
    
    # For procuradores, get their linked chefes
    procuradores = {}
    for u in all_users:
        if u.get('perfil') == 'Procurador':
            chefes_links = QueryBuilder("procurador_chefes").eq("id_procurador", u.get('id')).execute()
            chefe_ids = [c.get('id_chefe') for c in chefes_links]
            procuradores[u.get('id')] = {
                "nome": u.get('nome_completo'),
                "chefes": chefe_ids
            }
    
    return {"servidores": servidores, "chefes": chefes, "procuradores": procuradores}

def _format_value(value, is_percent=False):
    if pd.isna(value) or value is None: return 'Não Disponível'
    formatted_value = f"{value:.2f}" if isinstance(value, float) else str(value)
    return f"{formatted_value}%" if is_percent else formatted_value

def _calculate_average(series): return series.mean() if not series.empty else 0
def _calculate_percentage(series): return (series.sum() / len(series) * 100) if not series.empty else 0

def calcular_metricas_mensais(mes, ano):
    """Calcula métricas mensais usando Supabase API."""
    try:
        start_date = date(ano, mes, 1)
        end_date = (date(ano, mes, 1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        hierarchy = _get_user_hierarchy()
        
        # Fetch all processes and types
        all_processes = select_all("processos")
        all_product_types = select_all("tipos_produto")
        product_types_map = {p.get('id'): p for p in all_product_types}
        
        # Build DataFrame
        data = []
        for p in all_processes:
            tipo_produto = product_types_map.get(p.get('id_tipo_produto'), {})
            data.append({
                'id_servidor_responsavel': p.get('id_servidor_responsavel'),
                'id_chefe_gabinete': p.get('id_chefe_gabinete'),
                'id_procurador': p.get('id_procurador'),
                'data_atribuicao_servidor': p.get('data_atribuicao_servidor'),
                'data_conclusao_servidor': p.get('data_conclusao_servidor'),
                'data_conclusao_chefe': p.get('data_conclusao_chefe'),
                'prazo_servidor_aplicado': p.get('prazo_servidor_aplicado'),
                'prazo_chefe_aplicado': p.get('prazo_chefe_aplicado'),
                'prazo_total_dias_suspenso': p.get('prazo_total_dias_suspenso', 0),
                'status_servidor': p.get('status_servidor'),
                'status_chefe': p.get('status_chefe'),
                'tipo_contagem_prazo': tipo_produto.get('tipo_contagem_prazo', 'dias uteis')
            })
        
        df_full = pd.DataFrame(data)
        if df_full.empty:
            return {}

        # Data de corte para o relatório (último dia do mês)
        report_cutoff_dt = pd.to_datetime(end_date)

        for col in ['data_atribuicao_servidor', 'data_conclusao_servidor', 'data_conclusao_chefe']:
            df_full[col] = pd.to_datetime(df_full[col], errors='coerce')

        df_servidor_concluido = df_full.dropna(subset=['data_atribuicao_servidor', 'data_conclusao_servidor']).copy()
        df_servidor_concluido['duracao_servidor'] = df_servidor_concluido.apply(
            lambda row: utils.calculate_net_work_days(row['data_atribuicao_servidor'].date(), row['data_conclusao_servidor'].date(), row['id_servidor_responsavel']), axis=1
        )
        df_servidor_concluido['data_final_servidor'] = df_servidor_concluido.apply(
            lambda row: utils.calculate_due_date(row['data_atribuicao_servidor'].date(), row['prazo_servidor_aplicado'], row['tipo_contagem_prazo'], row['id_servidor_responsavel'], row['prazo_total_dias_suspenso']), axis=1
        )
        df_servidor_concluido['no_prazo_servidor'] = df_servidor_concluido['data_conclusao_servidor'].dt.date <= df_servidor_concluido['data_final_servidor']

        df_chefe_concluido = df_full.dropna(subset=['data_conclusao_servidor', 'data_conclusao_chefe']).copy()
        df_chefe_concluido['duracao_revisao_chefe'] = df_chefe_concluido.apply(
            lambda row: utils.calculate_net_work_days(row['data_conclusao_servidor'].date(), row['data_conclusao_chefe'].date(), row['id_chefe_gabinete']), axis=1
        )
        df_chefe_concluido['data_final_chefe'] = df_chefe_concluido.apply(
            lambda row: utils.calculate_due_date(row['data_conclusao_servidor'].date(), row['prazo_chefe_aplicado'], row['tipo_contagem_prazo'], row['id_chefe_gabinete'], row['prazo_total_dias_suspenso']), axis=1
        )
        df_chefe_concluido['no_prazo_chefe'] = df_chefe_concluido['data_conclusao_chefe'].dt.date <= df_chefe_concluido['data_final_chefe']
        
        df_servidor_mes = df_servidor_concluido[df_servidor_concluido['data_conclusao_servidor'].dt.date.between(start_date, end_date)]
        df_chefe_mes = df_chefe_concluido[df_chefe_concluido['data_conclusao_chefe'].dt.date.between(start_date, end_date)]
        
        m_servidor = {
            'avg_dias': df_servidor_mes.groupby('id_servidor_responsavel')['duracao_servidor'].mean(),
            'pct_prazo': df_servidor_mes.groupby('id_servidor_responsavel')['no_prazo_servidor'].apply(_calculate_percentage),
            # Acervo: processos não concluídos até o fim do mês, excluindo status finalizados (evita acervo fantasma)
            'acervo': df_full[
                df_full['data_conclusao_servidor'].isnull() & 
                (df_full['data_atribuicao_servidor'] <= report_cutoff_dt) & 
                (~df_full['status_servidor'].isin(['Concluído', 'Finalizado', 'Processo com o Procurador']))
            ].groupby('id_servidor_responsavel').size()
        }
        m_chefe = {
            'avg_dias_revisao': df_chefe_mes.groupby('id_chefe_gabinete')['duracao_revisao_chefe'].mean(),
            'pct_prazo_revisao': df_chefe_mes.groupby('id_chefe_gabinete')['no_prazo_chefe'].apply(_calculate_percentage),
            'num_revisados': df_chefe_mes.groupby('id_chefe_gabinete').size(),
            # Acervo revisão: processos concluídos pelo servidor mas não revisados até o fim do mês, excluindo status finalizados (evita acervo fantasma)
            'acervo_revisao': df_full[
                df_full['data_conclusao_servidor'].notnull() & 
                df_full['data_conclusao_chefe'].isnull() & 
                (df_full['data_conclusao_servidor'] <= report_cutoff_dt) & 
                (~df_full['status_chefe'].isin(['Finalizado', 'Processo com o Procurador']))
            ].groupby('id_chefe_gabinete').size()
        }

        metricas_finais = {}
        todos_procuradores = hierarchy['procuradores']

        metricas_finais["1) Média de dias que os pareceristas demoraram para concluir o processo (visão média por procurador)"] = {
            pdata['nome']: m_servidor['avg_dias'].reindex([sid for cid in pdata['chefes'] for sid in hierarchy['chefes'].get(cid, {}).get('servidores', [])]).mean() for _, pdata in todos_procuradores.items()
        }
        metricas_finais["2) Percentual de processos concluídos no prazo por pareceristas(visão média por procurador)"] = {
            pdata['nome']: m_servidor['pct_prazo'].reindex([sid for cid in pdata['chefes'] for sid in hierarchy['chefes'].get(cid, {}).get('servidores', [])]).mean() for _, pdata in todos_procuradores.items()
        }
        metricas_finais["3) Acervo de processo não concluídos ao encerrar o mês por parecerista (visão média por procurador)"] = {
            pdata['nome']: m_servidor['acervo'].reindex([sid for cid in pdata['chefes'] for sid in hierarchy['chefes'].get(cid, {}).get('servidores', [])]).sum() for _, pdata in todos_procuradores.items()
        }
        metricas_finais["4) Número de processos revisados no mês por chefe de gabinete (visão média por procurador)"] = {
            pdata['nome']: m_chefe['num_revisados'].reindex(pdata['chefes']).sum() for _, pdata in todos_procuradores.items()
        }
        metricas_finais["5) Média de dias que os chefes de gabinete demoraram para finalizar a revisão do processo (visão média por procurador)"] = {
            pdata['nome']: m_chefe['avg_dias_revisao'].reindex(pdata['chefes']).mean() for _, pdata in todos_procuradores.items()
        }
        metricas_finais["6) Percentual de processos revisados pelos chefes de gabinetes no prazo (visão média por procurador)"] = {
            pdata['nome']: m_chefe['pct_prazo_revisao'].reindex(pdata['chefes']).mean() for _, pdata in todos_procuradores.items()
        }
        metricas_finais["7) Acervo de processo não revisados ao encerrar o mês por chefe de gabinete (visão média por procurador)"] = {
            pdata['nome']: m_chefe['acervo_revisao'].reindex(pdata['chefes']).sum() for _, pdata in todos_procuradores.items()
        }

        return metricas_finais
        
    except Exception as e:
        import traceback
        print(f"ERRO DETALHADO em calcular_metricas_mensais: {e}\n{traceback.format_exc()}")
        return {}


def gerar_relatorio_xlsx(metricas, mes, ano):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Relatorio_{mes}_{ano}"
    font_bold = Font(bold=True)

    ws.cell(row=1, column=1, value="Métrica").font = font_bold
    ws.cell(row=1, column=2, value="Visão").font = font_bold
    ws.cell(row=1, column=3, value="Valor").font = font_bold
    
    row_idx = 2
    for metrica, visoes in sorted(metricas.items()):
        def format_value(v): return f"{v:.2f}" if isinstance(v, float) else v

        if isinstance(visoes, dict) and visoes:
            for visao, valor in visoes.items():
                ws.cell(row=row_idx, column=1, value=metrica)
                ws.cell(row=row_idx, column=2, value=str(visao))
                ws.cell(row=row_idx, column=3, value=format_value(valor))
                row_idx += 1
        else:
            ws.cell(row=row_idx, column=1, value=metrica)
            ws.cell(row=row_idx, column=3, value=format_value(visoes) if not isinstance(visoes, dict) else "N/A")
            row_idx += 1
            
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(cell.value)
            except: pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    filepath = f"relatorios/Relatorio_Produtividade_{mes}_{ano}.xlsx"
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    return filepath
